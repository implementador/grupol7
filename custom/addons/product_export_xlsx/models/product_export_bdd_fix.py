import base64
import io
import logging

from openpyxl import load_workbook

from odoo import models


_logger = logging.getLogger(__name__)


class ProductExportXlsxBDDFix(models.Model):
    _inherit = "product.export.xlsx"

    BDD_IVA_FACTOR = 1.16

    def _bdd_fix_find_pricelist(self, expected_name):
        Pricelist = (
            self.env["product.pricelist"]
            .sudo()
            .with_context(active_test=False)
        )

        exact = Pricelist.search(
            [
                ("name", "=", expected_name),
            ],
            limit=1,
        )

        if exact:
            return exact

        candidates = Pricelist.search(
            [
                ("name", "ilike", expected_name),
            ]
        )

        normalized = expected_name.strip().upper()

        for pricelist in candidates:
            if (pricelist.name or "").strip().upper() == normalized:
                return pricelist

        return candidates[:1]

    def _bdd_get_rule_types(self, pricelist, products):
        result = {}

        if not pricelist:
            return result

        PricelistItem = (
            self.env["product.pricelist.item"]
            .sudo()
            .with_context(active_test=False)
        )

        batch_size = 200

        for start in range(
            0,
            len(products),
            batch_size,
        ):
            batch = products[
                start:start + batch_size
            ]

            prices = pricelist._compute_price_rule(
                batch,
                1.0,
            )

            rule_ids = {
                value[1]
                for value in prices.values()
                if value
                and len(value) > 1
                and value[1]
            }

            rules = PricelistItem.browse(
                list(rule_ids)
            )

            applied_by_rule = {
                rule.id: rule.applied_on
                for rule in rules
            }

            for product in batch:
                value = prices.get(
                    product.id
                )

                rule_id = (
                    value[1]
                    if value
                    and len(value) > 1
                    else False
                )

                result[product.id] = (
                    applied_by_rule.get(
                        rule_id
                    )
                    if rule_id
                    else False
                )

        return result

    def _bdd_is_number(self, value):
        return (
            isinstance(value, (int, float))
            and not isinstance(value, bool)
        )

    def _bdd_gross(self, value, round_two=False):
        if not self._bdd_is_number(value):
            return "NA"

        gross = float(value) * self.BDD_IVA_FACTOR

        if round_two:
            return round(gross, 2)

        return gross

    def _bdd_fix_attachment(self):
        self.ensure_one()

        attachment = self.attachment_id.sudo()

        if not attachment:
            return

        if not attachment.datas:
            return

        raw = base64.b64decode(
            attachment.datas
        )

        workbook = load_workbook(
            io.BytesIO(raw)
        )

        if "BDD" not in workbook.sheetnames:
            _logger.warning(
                "Job %s: archivo principal sin hoja BDD",
                self.id,
            )
            return

        worksheet = workbook["BDD"]

        expected_headers = [
            "Referencia",
            "Producto",
            "Código de Barras",
            "Marca",
            "Categoria 1",
            "Categoria 2",
            "Categoria 3",
            "Categoria 4",
            "Condicionado",
            "Catalogo",
            "Categoría de UNSPSC",
            "Grupo",
            "Piezas",
            "UXE",
            "Imagen",
            "Costo",
            "Precio Intertienda",
            "Precio Mayoreo",
            "Precio Publico",
            "Precio Lista",
            "Precio Intertienda s/IVA",
            "Precio Mayoreo s/IVA",
            "Precio Publico s/IVA",
            "Precio Lista s/IVA",
        ]

        actual_headers = [
            worksheet.cell(
                1,
                col,
            ).value
            for col in range(
                1,
                25,
            )
        ]

        if actual_headers != expected_headers:
            raise ValueError(
                "Encabezados BDD no coinciden con formato esperado."
            )

        Product = (
            self.env["product.product"]
            .sudo()
            .with_context(active_test=False)
        )

        products = Product.search(
            [],
            order="id",
        )

        if worksheet.max_row - 1 != len(products):
            raise ValueError(
                "BDD y productos no tienen la misma cantidad. "
                "BDD=%s Productos=%s"
                % (
                    worksheet.max_row - 1,
                    len(products),
                )
            )

        intertienda = self._bdd_fix_find_pricelist(
            "PRECIO INTERTIENDA"
        )

        mayoreo = self._bdd_fix_find_pricelist(
            "PRECIO MAYOREO"
        )

        lista = self._bdd_fix_find_pricelist(
            "PRECIO LISTA"
        )

        _logger.info(
            "BDD V2.1 listas: Intertienda=%s Mayoreo=%s Lista=%s",
            intertienda.display_name if intertienda else "NO",
            mayoreo.display_name if mayoreo else "NO",
            lista.display_name if lista else "NO",
        )

        inter_rules = self._bdd_get_rule_types(
            intertienda,
            products,
        )

        may_rules = self._bdd_get_rule_types(
            mayoreo,
            products,
        )

        lista_rules = self._bdd_get_rule_types(
            lista,
            products,
        )

        # ------------------------------------------------------------------
        # UXE
        # El BDD historico utiliza principalmente:
        # product.template.x_studio_uxe_de_compra_2
        # En product.product existe como campo relacionado.
        # ------------------------------------------------------------------

        uxe_field = "x_studio_uxe_de_compra_2"

        uxe_by_product = {}

        if uxe_field in Product._fields:
            uxe_rows = products.read(
                [
                    "id",
                    uxe_field,
                ]
            )

            uxe_by_product = {
                row["id"]: row.get(uxe_field)
                for row in uxe_rows
            }

        stats = {
            "inter_global_na": 0,
            "may_global_na": 0,
            "lista_global_na": 0,
            "inter_numeric": 0,
            "may_numeric": 0,
            "lista_numeric": 0,
        }

        for index, product in enumerate(
            products,
            start=2,
        ):
            # --------------------------------------------------------------
            # UXE
            # --------------------------------------------------------------

            uxe_value = uxe_by_product.get(
                product.id
            )

            worksheet.cell(
                index,
                14,
            ).value = (
                uxe_value
                if uxe_value not in (
                    False,
                    None,
                    "",
                )
                else None
            )
            # --------------------------------------------------------------
            # CATEGORIAS BDD V2.1.4
            #
            # Replica la estructura del BDD historico:
            #
            # Todos / Vendibles / ACC PERSO / RUTA / CASCOS INF
            # ->
            # ACC PERSO / RUTA / RUTA / RUTA
            #
            # Todos / Vendibles / REFACC BIC
            # ->
            # Vendibles / Vendibles / Vendibles / Vendibles
            #
            # Todos / Vendibles / Mobility / Motobici
            # ->
            # Mobility / Mobility / Mobility / Mobility
            #
            # Se utiliza la jerarquia ACTUAL de Odoo.
            # --------------------------------------------------------------

            category = product.categ_id

            complete_name = (
                category.complete_name
                if category
                else ""
            )

            category_parts = [
                part.strip()
                for part in (
                    complete_name or ""
                ).split("/")
                if part.strip()
            ]

            # El BDD historico usa los padres de la categoria,
            # no la hoja final.
            if len(category_parts) > 1:
                category_parts = category_parts[:-1]

            # Si Todos es solamente raiz, eliminarlo.
            if (
                len(category_parts) > 1
                and category_parts[0].strip().lower() == "todos"
            ):
                category_parts = category_parts[1:]

            # Si Vendibles es solamente otra raiz y hay niveles
            # mas especificos, eliminarlo.
            if (
                len(category_parts) > 1
                and category_parts[0].strip().lower() == "vendibles"
            ):
                category_parts = category_parts[1:]

            if not category_parts and category:
                category_parts = [
                    category.name
                ]

            if category_parts:
                bdd_categories = category_parts[:4]

                while len(bdd_categories) < 4:
                    bdd_categories.append(
                        bdd_categories[-1]
                    )

            else:
                bdd_categories = [
                    None,
                    None,
                    None,
                    None,
                ]

            for category_offset, category_value in enumerate(
                bdd_categories,
                start=5,
            ):
                worksheet.cell(
                    index,
                    category_offset,
                ).value = category_value

            # --------------------------------------------------------------
            # Valores netos generados por V2
            # --------------------------------------------------------------

            net_cost = worksheet.cell(
                index,
                16,
            ).value

            net_inter = worksheet.cell(
                index,
                17,
            ).value

            net_may = worksheet.cell(
                index,
                18,
            ).value

            net_public = worksheet.cell(
                index,
                19,
            ).value

            net_lista = worksheet.cell(
                index,
                20,
            ).value

            # --------------------------------------------------------------
            # COSTO CON IVA
            # Referencia BDD:
            # standard_price 715 -> Costo 829.40
            # --------------------------------------------------------------

            if self._bdd_is_number(
                net_cost
            ):
                worksheet.cell(
                    index,
                    16,
                ).value = round(
                    float(net_cost)
                    * self.BDD_IVA_FACTOR,
                    2,
                )

            # --------------------------------------------------------------
            # INTERTIENDA
            # Las reglas globales no se muestran en el BDD original.
            # --------------------------------------------------------------

            inter_rule_type = inter_rules.get(
                product.id
            )

            inter_valid = (
                self._bdd_is_number(
                    net_inter
                )
                and inter_rule_type
                and inter_rule_type != "3_global"
            )

            if inter_valid:
                worksheet.cell(
                    index,
                    17,
                ).value = (
                    float(net_inter)
                    * self.BDD_IVA_FACTOR
                )

                worksheet.cell(
                    index,
                    21,
                ).value = float(
                    net_inter
                )

                stats[
                    "inter_numeric"
                ] += 1

            else:
                worksheet.cell(
                    index,
                    17,
                ).value = "NA"

                worksheet.cell(
                    index,
                    21,
                ).value = 0

                if (
                    inter_rule_type
                    == "3_global"
                ):
                    stats[
                        "inter_global_na"
                    ] += 1

            # --------------------------------------------------------------
            # MAYOREO
            # --------------------------------------------------------------

            may_rule_type = may_rules.get(
                product.id
            )

            may_valid = (
                self._bdd_is_number(
                    net_may
                )
                and may_rule_type
                and may_rule_type != "3_global"
            )

            if may_valid:
                worksheet.cell(
                    index,
                    18,
                ).value = (
                    float(net_may)
                    * self.BDD_IVA_FACTOR
                )

                worksheet.cell(
                    index,
                    22,
                ).value = float(
                    net_may
                )

                stats[
                    "may_numeric"
                ] += 1

            else:
                worksheet.cell(
                    index,
                    18,
                ).value = "NA"

                worksheet.cell(
                    index,
                    22,
                ).value = 0

                if (
                    may_rule_type
                    == "3_global"
                ):
                    stats[
                        "may_global_na"
                    ] += 1

            # --------------------------------------------------------------
            # PRECIO PUBLICO
            # BDD original redondea precio con IVA a 2 decimales.
            # --------------------------------------------------------------

            if self._bdd_is_number(
                net_public
            ):
                worksheet.cell(
                    index,
                    19,
                ).value = round(
                    float(net_public)
                    * self.BDD_IVA_FACTOR,
                    2,
                )

                worksheet.cell(
                    index,
                    23,
                ).value = round(
                    float(net_public),
                    2,
                )

            else:
                worksheet.cell(
                    index,
                    19,
                ).value = 0

                worksheet.cell(
                    index,
                    23,
                ).value = 0

            # --------------------------------------------------------------
            # PRECIO LISTA
            # --------------------------------------------------------------

            lista_rule_type = lista_rules.get(
                product.id
            )

            lista_valid = (
                self._bdd_is_number(
                    net_lista
                )
                and lista_rule_type
                and lista_rule_type != "3_global"
            )

            if lista_valid:
                worksheet.cell(
                    index,
                    20,
                ).value = (
                    float(net_lista)
                    * self.BDD_IVA_FACTOR
                )

                worksheet.cell(
                    index,
                    24,
                ).value = float(
                    net_lista
                )

                stats[
                    "lista_numeric"
                ] += 1

            else:
                worksheet.cell(
                    index,
                    20,
                ).value = "NA"

                worksheet.cell(
                    index,
                    24,
                ).value = 0

                if (
                    lista_rule_type
                    == "3_global"
                ):
                    stats[
                        "lista_global_na"
                    ] += 1

        # ------------------------------------------------------------------
        # Igualar dimensiones al BDD usado en Documentos
        # ------------------------------------------------------------------

        worksheet.freeze_panes = "A2"

        worksheet.column_dimensions[
            "A"
        ].width = 13

        worksheet.column_dimensions[
            "B"
        ].width = 57.140625

        for col in [
            "C", "D", "E", "F", "G", "H",
            "I", "J", "K", "L", "M", "N",
            "O", "P", "Q", "R", "S", "T",
            "U", "V", "W", "X",
        ]:
            worksheet.column_dimensions[
                col
            ].width = 13

        output = io.BytesIO()

        workbook.save(
            output
        )

        binary = output.getvalue()

        attachment.write({
            "datas": base64.b64encode(
                binary
            ),
        })

        old_progress = (
            self.progress_message
            or ""
        )

        self.write({
            "file_size_mb": (
                len(binary)
                / 1024
                / 1024
            ),
            "progress_message": (
                "BDD V2.1 lista | "
                "IVA 16%% corregido | "
                "Intertienda global->NA: %s | "
                "Mayoreo global->NA: %s | "
                "Lista global->NA: %s | %s"
                % (
                    stats[
                        "inter_global_na"
                    ],
                    stats[
                        "may_global_na"
                    ],
                    stats[
                        "lista_global_na"
                    ],
                    old_progress,
                )
            ),
        })

        _logger.info(
            "BDD V2.1 Job %s stats: %s",
            self.id,
            stats,
        )

    def _generate_xlsx(self):
        result = super()._generate_xlsx()

        for record in self:
            if (
                record.state == "done"
                and record.attachment_id
            ):
                record._bdd_fix_attachment()

        return result
